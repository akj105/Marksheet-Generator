import csv
import openpyxl
from openpyxl.styles import Font
import os
os.system('cls')
from pywebio.input import *
from pywebio.output import *
from pywebio import start_server
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders
   
fromaddr = "aj5333789"
passwrd ="gmailaccount"

def validity(file):
    content_1 = file['content'].decode('utf-8').splitlines()
    rdr= csv.reader(content_1)
    for arr in rdr:
        if arr[6]=="ANSWER":
            i=0
            j=i-6
            for num in arr:
                if len(num)!=0 and i>6:
                    dict1[j]=num
                i=i+1
                j=j+1

            return
    
    put_text('no roll number with ANSWER is present, Cannot Process!')
    response = file_upload("Browse for response CSV", accept=".csv")
    validity(response)


def generate_marksheet(line):
    path=r"./marksheets/"
    roll=line[0]
    name=line[1]
    content_3 = response['content'].decode('utf-8').splitlines()
    c_rr=csv.reader(content_3)
    title2=next(c_rr)
    for ar in c_rr:
        if ar[6].upper()==roll:
            i=0
            j=i-6
            for num in ar:
                if i>6 and i<=i+len(dict1):
                    dict2[j]=num
                i=i+1
                j=j+1

            wb = openpyxl.Workbook()
            sheet=wb.active
            sheet['A18']="Student Ans"
            sheet['A18'].font=Font(bold=True)
            sheet['B18']="Correct Ans"
            sheet['B18'].font=Font(bold=True)
            sheet['A9']="Name:"
            sheet['B9']=name
            sheet['B9'].font=Font(bold=True)
            sheet['A10']="Roll no:"
            sheet['B10']=roll
            sheet['B10'].font=Font(bold=True)
            sheet['D9']="Exam:"
            sheet['E9']="quiz"
            sheet['E9'].font=Font(bold=True)
            sheet['B12']="Right"
            sheet['B12'].font=Font(bold=True)
            sheet['C12']="Wrong"
            sheet['C12'].font=Font(bold=True)
            sheet['D12']="Not Attempt"
            sheet['D12'].font=Font(bold=True)
            sheet['E12']="Max"
            sheet['E12'].font=Font(bold=True)
            sheet['A13']="No."
            sheet['A13'].font=Font(bold=True)
            sheet['A14']="Marking"
            sheet['A14'].font=Font(bold=True)
            sheet['A15']="Total"
            sheet['A15'].font=Font(bold=True)
            sheet['E13']=len(dict1)
            sheet['B14']=positive
            sheet['C14']=negative
            sheet['D14']=0
            x=0
            y=0
            img = openpyxl.drawing.image.Image('pic1.png')
            img.anchor = 'A1'
            sheet.add_image(img)
            for p in dict1:
                sheet.cell(row=int(p+18),column=2).value=dict1[p]
                sheet.cell(row=int(p+18),column=2).font=Font(color="E60000FF")
                if(len(dict2[p])==0):
                    y=y+1
                if(dict1[p]==dict2[p]):
                    sheet.cell(row=int(p+18),column=1).value=dict2[p]
                    sheet.cell(row=int(p+18),column=1).font=Font(color="E6008000")
                    x=x+1
                else:
                    sheet.cell(row=int(p+18),column=1).value=dict2[p]
                    sheet.cell(row=int(p+18),column=1).font=Font(color="E6FF0000")
                    

            sheet['B13']=x
            sheet['C13']=len(dict1)-(x+y)
            sheet['D13']=y
            sheet['B15']=x*positive
            sheet['C15']=(len(dict1)-(x+y))*negative
            sheet['E15']=str(x*positive+(len(dict1)-(x+y))*negative)+"/"+str(len(dict1)*positive)
            wb.save(path+'%s.xlsx' %ar[6].upper())
            with open("./marksheets/concise_marksheet.csv", "a") as file:
                for detail in ar:
                    file.write(detail)
                    file.write(',')
                file.write(str(x*positive+(len(dict1)-(x+y))*negative)+"/"+str(len(dict1)*positive))
                file.write("\n")
            return

    i=0
    j=i-6
    while(1):
        if i>6 and i<=i+len(dict1):
            dict2[j]=0
        i=i+1
        j=j+1

    wb = openpyxl.Workbook()
    sheet=wb.active
    sheet['A18']="Student Ans"
    sheet['A18'].font=Font(bold=True)
    sheet['B18']="Correct Ans"
    sheet['B18'].font=Font(bold=True)
    sheet['A9']="Name:"
    sheet['B9']=name
    sheet['B9'].font=Font(bold=True)
    sheet['A10']="Roll no:"
    sheet['B10']=roll
    sheet['B10'].font=Font(bold=True)
    sheet['D9']="Exam:"
    sheet['E9']="quiz"
    sheet['E9'].font=Font(bold=True)
    sheet['B12']="Right"
    sheet['B12'].font=Font(bold=True)
    sheet['C12']="Wrong"
    sheet['C12'].font=Font(bold=True)
    sheet['D12']="Not Attempt"
    sheet['D12'].font=Font(bold=True)
    sheet['E12']="Max"
    sheet['E12'].font=Font(bold=True)
    sheet['A13']="No."
    sheet['A13'].font=Font(bold=True)
    sheet['A14']="Marking"
    sheet['A14'].font=Font(bold=True)
    sheet['A15']="Total"
    sheet['A15'].font=Font(bold=True)
    sheet['E13']=len(dict1)
    sheet['B14']=positive
    sheet['C14']=negative
    sheet['D14']=0
    x=0
    y=0
    img = openpyxl.drawing.image.Image('pic1.png')
    img.anchor = 'A1'
    sheet.add_image(img)
    for p in dict1:
        sheet.cell(row=int(p+18),column=2).value=dict1[p]
        sheet.cell(row=int(p+18),column=2).font=Font(color="E60000FF")
        if(len(dict2[p])==0):
            y=y+1
        if(dict1[p]==dict2[p]):
            sheet.cell(row=int(p+18),column=1).value=dict2[p]
            sheet.cell(row=int(p+18),column=1).font=Font(color="E6008000")
            x=x+1
        else:
            sheet.cell(row=int(p+18),column=1).value=dict2[p]
            sheet.cell(row=int(p+18),column=1).font=Font(color="E6FF0000")
                    

    sheet['B13']=x
    sheet['C13']=len(dict1)-(x+y)
    sheet['D13']=y
    sheet['B15']=x*positive
    sheet['C15']=(len(dict1)-(x+y))*negative
    sheet['E15']=str(x*positive+(len(dict1)-(x+y))*negative)+"/"+str(len(dict1)*positive)
    wb.save(path+'%s.xlsx' %roll)

    return       

def send_email(line):
    toaddr = line[1]
    roll=line[6]
    roll=roll.upper()
    msg = MIMEMultipart()
   
    msg['From'] = fromaddr 
    msg['To'] = toaddr
  
    msg['Subject'] = "QUIZ MARKSHEET"
  
    body = "Please find attached marksheet of your quiz"
  
    msg.attach(MIMEText(body, 'plain'))
  
    # opening the file to be sent 
    filename = "%s.xlsx"%roll
    attachment = open("./marksheets/%s.xlsx"%roll, "rb")
  
    p = MIMEBase('application', 'octet-stream')
  
    p.set_payload((attachment).read())
  
    encoders.encode_base64(p)
   
    p.add_header('Content-Disposition', "attachment; filename= %s" % filename)
  
    msg.attach(p)

    s = smtplib.SMTP('smtp.gmail.com', 587)
  
    s.starttls()
  
    s.login(fromaddr, passwrd)
  
    text = msg.as_string()
    try:
        s.sendmail(fromaddr, toaddr, text)
    except:
        print("Email ID format is wrong")
    
    s.quit()

    #sending to 2nd email

    toaddr = line[4]
    roll=line[6]
    roll=roll.upper()
    msg = MIMEMultipart()
   
    msg['From'] = fromaddr 
    msg['To'] = toaddr
  
    msg['Subject'] = "QUIZ MARKSHEET"
  
    body = "Please find attached marksheet of your quiz"
  
    msg.attach(MIMEText(body, 'plain'))
  
    # opening the file to be sent 
    filename = "%s.xlsx"%roll
    attachment = open("./marksheets/%s.xlsx"%roll, "rb")
  
    p = MIMEBase('application', 'octet-stream')
  
    p.set_payload((attachment).read())
  
    encoders.encode_base64(p)
   
    p.add_header('Content-Disposition', "attachment; filename= %s" % filename)
  
    msg.attach(p)

    s = smtplib.SMTP('smtp.gmail.com', 587)
  
    s.starttls()
  
    s.login(fromaddr, passwrd)
  
    text = msg.as_string()

    try:
        s.sendmail(fromaddr, toaddr, text)
    except:
        print("Email ID format is wrong")
    s.quit()


    return

dict1={}
dict2={}
master_roll = file_upload("Browse for master_roll CSV", accept=".csv",required=True)
response = file_upload("Browse for response CSV", accept=".csv")
validity(response)


positive = input("Marks for correct answer", type=FLOAT)
negative = input("-ve marks for wrong answer", type=FLOAT)

main_dir="./marksheets"
if os.path.exists(main_dir) == False:
    os.mkdir(main_dir)

with open("./marksheets/concise_marksheet.csv", "a") as file:
    file.write("Timestamp,Email_address,Google_Score,Name,IITP_webmail,Phone,Roll_Number,Q1,Q2,Q3,Q4,Q5,Q6,Q7,Q8,Q9,Q10,Q11,Q12,Q13,Q14,Q15,Q16,Q17,Q18,Q19,Q20,Q21,Q22,Q23,Q24,Q25,Q26,Q27,Q28,Score_After_Negative\n")

put_text('Generate Marksheet \n')
mark_s = select('Select option:', ["YES", "NO"])

if mark_s == "YES":
    content_2 = master_roll['content'].decode('utf-8').splitlines()
    c_r=csv.reader(content_2)
    title1=next(c_r)
    for arr in c_r:
        generate_marksheet(arr)


put_text('Send Email \n')
mail_s = select('Select option:', ["YES", "NO"])

if mail_s == "YES":
    content_4 = response['content'].decode('utf-8').splitlines()
    c_rs=csv.reader(content_4)
    title2=next(c_rs)
    for arr in c_rs:
        send_email(arr)